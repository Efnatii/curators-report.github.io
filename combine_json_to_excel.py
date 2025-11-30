"""
Combine multiple JSON survey response files into a single Excel workbook.

Usage (CLI):
    python combine_json_to_excel.py /path/to/json_dir output.xlsx

If the script is launched without CLI arguments, a simple form-style GUI will
open to select the source folder and output file path.

Each JSON file should contain a flat mapping of question text to either a
scalar answer (string/number) or a list of strings (representing sub-rows).
When a question contains a list, the corresponding cells in Excel will be
highlighted to distinguish the block of rows taken by that subtable.
"""

from __future__ import annotations

import argparse
import importlib.util
from dataclasses import dataclass
import json
import base64
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import subprocess
import sys
from typing import Dict, Iterable, List, Tuple

from icon_data import ICON_BASE64


REQUIRED_PACKAGES = ["openpyxl", "fpdf2"]


def ensure_dependencies_installed() -> None:
    """Install required third-party packages at runtime when missing."""

    for package in REQUIRED_PACKAGES:
        if importlib.util.find_spec(package) is None:
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            except subprocess.CalledProcessError as exc:
                raise RuntimeError(
                    f"Не удалось автоматически установить пакет {package}. "
                    "Убедитесь, что есть доступ в интернет или установите пакет вручную."
                ) from exc


ensure_dependencies_installed()

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF


def get_resource_path(relative_path: str) -> Path:
    """Return an absolute path to a bundled resource (PyInstaller friendly)."""

    base_path = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base_path / relative_path


def ensure_icon_path() -> Path:
    """Make sure the ICO file exists locally when not running from a bundle."""

    icon_path = get_resource_path("combine_json_to_excel.ico")
    if icon_path.exists():
        return icon_path

    if hasattr(sys, "_MEIPASS"):
        # In a bundled executable the icon should already be present alongside the binary.
        return icon_path

    try:
        icon_bytes = base64.b64decode("".join(ICON_BASE64))
        icon_path.write_bytes(icon_bytes)
    except OSError:
        # If the path is not writable, silently continue without the icon.
        pass
    return icon_path


ListValue = List[str]
ScalarValue = str | int | float | bool | None
JsonValue = ScalarValue | ListValue
JsonRecord = Dict[str, JsonValue]


HIGHLIGHT_FILL = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")


def format_scalar_list(items: ListValue) -> str:
    """Sort list items and join them with a comma for display."""

    sorted_items = sorted((str(item) for item in items), key=str)
    return ", ".join(sorted_items)


def normalize_cell_value(value: JsonValue) -> ScalarValue:
    """Convert any JSON value into something openpyxl can store."""

    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, list) and all(not isinstance(item, (dict, list)) for item in value):
        return format_scalar_list(value)
    # Fallback for types like dicts or lists of non-scalar values
    return json.dumps(value, ensure_ascii=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge JSON files into a single Excel workbook where each question becomes a column.",
    )
    parser.add_argument(
        "input_dir",
        type=Path,
        help="Directory containing JSON files to merge.",
    )
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=Path("combined.xlsx"),
        help="Path for the generated Excel file (default: combined.xlsx).",
    )
    parser.add_argument(
        "--generate-pdf",
        action="store_true",
        help="Generate per-respondent PDF с подробной разбивкой баллов.",
    )
    return parser.parse_args()


def load_json_files(input_dir: Path) -> List[Tuple[Path, JsonRecord]]:
    json_files = sorted(input_dir.glob("*.json"))
    if not json_files:
        raise FileNotFoundError(f"No JSON files found in {input_dir}")

    records: List[Tuple[Path, JsonRecord]] = []
    for path in json_files:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            raise ValueError(f"JSON file {path} does not contain an object at the top level.")
        records.append((path, data))
    return records


QUESTION_ORDER: List[Dict[str, object]] = [
    {
        "key": "reporting_period",
        "label": "Отчётный период",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
        },
    },
    {"key": "score", "label": "Баллы", "always": True},
    {"key": "full_name", "label": "1. Фамилия Имя Отчество"},
    {"key": "job_positions", "label": "2. Должность"},
    {"key": "department", "label": "3. Кафедра"},
    {
        "key": "contact_phone_connected_to_telegram",
        "label": "4. Контактный телефон (подключенный к Telegram)",
    },
    {"key": "telegram_username", "label": "5. Ник в Telegram"},
    {"key": "email", "label": "6. E-mail"},
    {"key": "curated_group_numbers", "label": "7. Номера курируемых групп"},
    {"key": "curator_primary_building", "label": "8. Корпус основного пребывания куратора"},
    {"key": "curator_primary_room", "label": "9. Аудитория основного пребывания куратора"},
    {"key": "institute_or_faculty", "label": "10. Институт/Факультет"},
    {
        "key": "held_minimum_three_curator_sessions_in_reporting_period",
        "label": "11. Проведение не менее трёх кураторских часов за отчётный период",
    },
    {
        "key": "curator_hours_details",
        "label": "12. Даты проведения трёх и более кураторских часов в течение отчётного периода",
        "subfields": {
            "groups": "Группы",
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "topic": "Тема",
            "directions": "Направленность",
            "specialists": "Приглашённые специалисты",
        },
    },
    {"key": "manages_group_chat", "label": "13. Ведение чата с каждой группой или общего чата"},
    {
        "key": "inform_group_about_events",
        "label": "14. Информирование группы о мероприятиях и событиях различного уровня",
    },
    {
        "key": "achievements",
        "label": "15. Призовое место обучающегося во внеучебных мероприятиях",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
            "group": "Группа",
            "student": "ФИО студента",
            "result": "Итог",
        },
    },
    {
        "key": "participated_in_two_events_with_group",
        "label": "16. Совместное участие с группой не менее чем в двух мероприятиях",
    },
    {
        "key": "joint_participation_events",
        "label": "17. Даты проведения двух и более мероприятий в течение отчётного периода",
        "subfields": {
            "groups": "Группы",
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
        },
    },
    {
        "key": "participated_in_two_curator_events",
        "label": "18. Участие не менее чем в двух мероприятиях для кураторов",
    },
    {
        "key": "curator_personal_events",
        "label": "19. Даты участия в мероприятиях для кураторов",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
        },
    },
    {
        "key": "personal_program_participation",
        "label": "20. Личное участие в программах и конкурсах",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
        },
    },
    {
        "key": "mentor_support_events",
        "label": "21. Участие куратора в роли наставника проекта",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
            "group": "Группа",
            "student": "ФИО студента",
            "result": "Итог",
        },
    },
    {
        "key": "scientific_publications",
        "label": "22. Опубликование научной работы",
        "subfields": {
            "description": "Описание",
            "link": "Ссылка",
        },
    },
    {
        "key": "media_materials",
        "label": "23. Интервью и статьи для \"Дзен.Гуап\", соцсетей и сайта ГУАП",
        "subfields": {"link": "Ссылка"},
    },
    {
        "key": "qualification_courses",
        "label": "24. Курсы повышения квалификации",
        "subfields": {
            "date_start": "Дата начала",
            "date_end": "Дата окончания",
            "event": "Мероприятие",
        },
    },
]


@dataclass(frozen=True)
class QuestionColumn:
    key: str
    label: str
    subfields: List[Tuple[str, str]] | None = None


def determine_columns(records: Iterable[Tuple[Path, JsonRecord]]) -> List[QuestionColumn]:
    present_keys = {key for _, record in records for key in record.keys()}
    columns: List[QuestionColumn] = []

    for question in QUESTION_ORDER:
        key = question["key"]  # type: ignore[index]
        always_include = bool(question.get("always"))
        if not always_include and key not in present_keys:
            continue

        label = question["label"]  # type: ignore[index]
        subfields: Dict[str, str] | None = question.get("subfields")  # type: ignore[assignment]

        if subfields:
            columns.append(QuestionColumn(key=key, label=label, subfields=list(subfields.items())))
        else:
            columns.append(QuestionColumn(key=key, label=label))

    return columns


def normalize_reporting_period(value: JsonValue) -> Dict[str, ScalarValue]:
    """Drop legacy keys and split the range into explicit dates."""

    date_start: ScalarValue = None
    date_end: ScalarValue = None

    if isinstance(value, dict):
        date_start = value.get("date_start")  # type: ignore[assignment]
        date_end = value.get("date_end")  # type: ignore[assignment]

        if date_start is None and date_end is None:
            range_value = value.get("range")
            if isinstance(range_value, str):
                start, _, end = range_value.partition(" - ")
                if start and end:
                    date_start = start
                    date_end = end

    return {"date_start": date_start, "date_end": date_end}


def normalize_record_values(record: JsonRecord) -> JsonRecord:
    """Merge scalar lists into comma-separated strings for single-cell output."""

    normalized: JsonRecord = {}
    for key, value in record.items():
        if key == "reporting_period":
            normalized[key] = normalize_reporting_period(value)
            continue

        if isinstance(value, list) and all(not isinstance(item, (dict, list)) for item in value):
            normalized[key] = format_scalar_list(value)
        else:
            normalized[key] = value

    return normalized


def is_yes(value: JsonValue) -> bool:
    """Return True when the answer represents an affirmative response."""

    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() == "да"
    return False


def row_has_content(row: JsonValue) -> bool:
    """Check whether a list item should be counted as a filled row."""

    if isinstance(row, dict):
        return any(bool(v) for v in row.values())
    if isinstance(row, (str, int, float, bool)):
        return bool(row)
    return False


def count_filled_rows(value: JsonValue) -> int:
    """Count non-empty rows inside a list value."""

    if not isinstance(value, list):
        return 0
    return sum(1 for item in value if row_has_content(item))


def count_rows_with_specialists(value: JsonValue) -> int:
    """Count curator hour rows that include invited specialists."""

    if not isinstance(value, list):
        return 0

    count = 0
    for item in value:
        if not isinstance(item, dict):
            continue
        specialists = item.get("specialists")
        if isinstance(specialists, str) and specialists.strip():
            count += 1
        elif isinstance(specialists, list) and any(specialist for specialist in specialists):
            count += 1
    return count


def compute_point_components(record: JsonRecord) -> List[Tuple[str, int]]:
    """Return detailed point components for a survey record."""

    value_11 = record.get("held_minimum_three_curator_sessions_in_reporting_period")
    value_12 = record.get("curator_hours_details")
    value_13 = record.get("manages_group_chat")
    value_14 = record.get("inform_group_about_events")
    value_16 = record.get("participated_in_two_events_with_group")
    value_17 = record.get("joint_participation_events")
    value_18 = record.get("participated_in_two_curator_events")
    value_19 = record.get("curator_personal_events")

    count_12 = count_filled_rows(value_12)
    count_17 = count_filled_rows(value_17)
    count_19 = count_filled_rows(value_19)

    components: List[Tuple[str, int]] = []

    base_condition_met = (
        is_yes(value_11)
        and count_12 >= 3
        and is_yes(value_13)
        and is_yes(value_14)
        and is_yes(value_16)
        and count_17 >= 2
        and is_yes(value_18)
        and count_19 >= 2
    )

    components.append(
        (
            "Базовое условие (пп. 11, 12, 13, 14, 16, 17, 18, 19)",
            30 if base_condition_met else 0,
        )
    )

    components.append(
        (
            "Личное участие в программах и конкурсах (п.20)",
            count_filled_rows(record.get("personal_program_participation")) * 10,
        )
    )
    components.append(
        (
            "Опубликование научной работы (п.22)",
            count_filled_rows(record.get("scientific_publications")) * 20,
        )
    )
    components.append(
        (
            "Интервью и статьи для \"Дзен.Гуап\" и соцсетей (п.23)",
            count_filled_rows(record.get("media_materials")) * 10,
        )
    )
    components.append(
        (
            "Наставничество в проектах (п.21)",
            count_filled_rows(record.get("mentor_support_events")) * 10,
        )
    )
    components.append(
        (
            "Призовые места обучающихся (п.15)",
            count_filled_rows(record.get("achievements")) * 10,
        )
    )
    components.append(
        (
            "Курсы повышения квалификации (п.24)",
            count_filled_rows(record.get("qualification_courses")) * 20,
        )
    )

    components.append(
        (
            "Приглашённые специалисты на кураторских часах (п.12)",
            count_rows_with_specialists(value_12) * 20,
        )
    )

    components.append(
        (
            "Дополнительные кураторские часы сверх трёх (п.12)",
            (count_12 - 3) * 5 if count_12 > 3 else 0,
        )
    )

    components.append(
        (
            "Совместные мероприятия с группой сверх двух (п.17)",
            (count_17 - 2) * 10 if count_17 > 2 else 0,
        )
    )

    components.append(
        (
            "Мероприятия для кураторов сверх двух (п.19)",
            (count_19 - 2) * 5 if count_19 > 2 else 0,
        )
    )

    return components


def compute_points(record: JsonRecord) -> int:
    """Calculate the total score for a single survey record."""

    return sum(points for _, points in compute_point_components(record))


def stringify_value(value: JsonValue) -> str:
    if value is None:
        return ""
    if isinstance(value, list) and all(not isinstance(item, (dict, list)) for item in value):
        return format_scalar_list(value)
    return str(value)


def sanitize_filename(value: str) -> str:
    safe_value = re.sub(r"[\\/:*?\"<>|]", "_", value).strip()
    return safe_value or "Без_ФИО"


def extract_full_name(record: JsonRecord, fallback: str) -> str:
    full_name_value = record.get("full_name")
    full_name = stringify_value(full_name_value)
    return full_name if full_name else fallback


def render_pdf_row(pdf: FPDF, text: str, points: int, column_width: float, points_width: float) -> None:
    """Render a wrapped table row with aligned points column."""

    line_height = 8
    x_start = pdf.get_x()
    y_start = pdf.get_y()

    pdf.multi_cell(column_width, line_height, text, border=1)
    y_end = pdf.get_y()
    row_height = y_end - y_start

    pdf.set_xy(x_start + column_width, y_start)
    pdf.cell(points_width, row_height, str(points), border=1, align="R")
    pdf.set_xy(x_start, y_end)


def generate_score_pdf(
    *,
    full_name: str,
    components: List[Tuple[str, int]],
    output_path: Path,
    total: int,
) -> None:
    pdf = FPDF()
    pdf.add_page()

    font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    if font_path.exists():
        pdf.add_font("DejaVu", "", str(font_path), uni=True)
        pdf.add_font("DejaVu", "B", str(font_path), uni=True)
        body_font = ("DejaVu", "")
        heading_font = ("DejaVu", "B")
    else:
        body_font = ("Arial", "")
        heading_font = ("Arial", "B")

    pdf.set_font(*heading_font, size=14)
    pdf.cell(0, 10, "Отчёт по баллам", ln=True)

    pdf.set_font(*body_font, size=12)
    pdf.cell(0, 10, f"ФИО: {full_name}", ln=True)
    pdf.ln(2)

    table_width = pdf.w - pdf.l_margin - pdf.r_margin
    points_width = 30
    column_width = table_width - points_width

    pdf.set_font(*heading_font, size=11)
    pdf.cell(column_width, 8, "Критерий", border=1)
    pdf.cell(points_width, 8, "Баллы", border=1, ln=True, align="R")

    pdf.set_font(*body_font, size=11)
    for description, points in components:
        render_pdf_row(pdf, description, points, column_width, points_width)

    pdf.set_font(*heading_font, size=12)
    pdf.cell(column_width, 10, "Итого", border=1)
    pdf.cell(points_width, 10, str(total), border=1, ln=True, align="R")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(output_path))


def generate_score_reports(records: List[Tuple[Path, JsonRecord]], target_dir: Path) -> None:
    for file_path, raw_record in records:
        components = compute_point_components(raw_record)
        total_points = sum(points for _, points in components)
        full_name = extract_full_name(raw_record, fallback=file_path.stem)
        safe_name = sanitize_filename(full_name)
        pdf_name = f"Баллы_{safe_name}.pdf"
        pdf_path = target_dir / pdf_name
        generate_score_pdf(full_name=full_name, components=components, output_path=pdf_path, total=total_points)


def write_workbook(
    records: List[Tuple[Path, JsonRecord]],
    questions: List[QuestionColumn],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    ws.cell(row=1, column=1, value="Источник")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    flat_columns: List[Tuple[str, str | None]] = []
    column_index = 2
    for question in questions:
        span = len(question.subfields) if question.subfields else 1

        if question.subfields:
            ws.merge_cells(
                start_row=1,
                start_column=column_index,
                end_row=1,
                end_column=column_index + span - 1,
            )
            ws.cell(row=1, column=column_index, value=question.label)
            for offset, (subkey, sublabel) in enumerate(question.subfields):
                ws.cell(row=2, column=column_index + offset, value=sublabel)
                flat_columns.append((question.key, subkey))
        else:
            ws.merge_cells(
                start_row=1,
                start_column=column_index,
                end_row=2,
                end_column=column_index,
            )
            ws.cell(row=1, column=column_index, value=question.label)
            flat_columns.append((question.key, None))

        column_index += span

    alignment_top = Alignment(vertical="top")
    for header_cell in ws[1] + ws[2]:
        if header_cell.value is not None:
            header_cell.alignment = alignment_top

    for file_path, raw_record in records:
        record = normalize_record_values(raw_record)
        record["score"] = compute_points(raw_record)
        block_height = max(
            (
                len(value) if isinstance(value, list) else 1
                for value in record.values()
            ),
            default=1,
        )
        for row_offset in range(block_height):
            row_values = [file_path.name if row_offset == 0 else ""]
            for question_key, subkey in flat_columns:
                value = record.get(question_key)
                if isinstance(value, list):
                    if row_offset < len(value):
                        item = value[row_offset]
                        if subkey is not None and isinstance(item, dict):
                            cell_value = normalize_cell_value(item.get(subkey))
                        elif subkey is None:
                            cell_value = normalize_cell_value(item)
                        else:
                            cell_value = ""
                    else:
                        cell_value = ""
                elif isinstance(value, dict):
                    if subkey is not None:
                        cell_value = normalize_cell_value(value.get(subkey)) if row_offset == 0 else ""
                    elif row_offset == 0:
                        cell_value = normalize_cell_value(value)
                    else:
                        cell_value = ""
                else:
                    if subkey is not None:
                        cell_value = ""
                    else:
                        cell_value = normalize_cell_value(value) if row_offset == 0 else ""
                row_values.append(cell_value)
            ws.append(row_values)

        # Apply highlighting to list ranges per question
        start_row = ws.max_row - block_height + 1
        column_start = 2
        for question in questions:
            span = len(question.subfields) if question.subfields else 1
            value = record.get(question.key)
            if isinstance(value, list) and value:
                for offset in range(span):
                    for row_index in range(start_row, start_row + len(value)):
                        ws.cell(row=row_index, column=column_start + offset).fill = HIGHLIGHT_FILL
            column_start += span

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).split("\n"):
                max_length = max(max_length, len(line))

        if max_length:
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_path)


def merge_json_directory(input_dir: Path, output_path: Path, generate_pdfs: bool = False) -> int:
    """Merge JSON files from a directory into an Excel workbook."""

    records = load_json_files(input_dir)
    questions = determine_columns(records)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(records, questions, output_path)
    if generate_pdfs:
        generate_score_reports(records, output_path.parent)
    return len(records)


def launch_gui() -> None:
    """Open a simple form-style interface for combining survey exports."""

    root = tk.Tk()
    root.title("Сбор анкет в Excel")
    root.configure(bg="#0c0f16")
    root.resizable(False, False)

    icon_path = ensure_icon_path()
    if icon_path.exists():
        try:
            root.iconbitmap(default=str(icon_path))
        except tk.TclError:
            try:
                icon_image = tk.PhotoImage(file=str(icon_path))
                root.iconphoto(False, icon_image)
                root._icon_image = icon_image  # prevent garbage collection
            except tk.TclError:
                pass

    palette = {
        "bg": "#0c0f16",
        "panel": "#0f1522",
        "card": "#111a2b",
        "border": "#1f2a3d",
        "accent": "#f2b138",
        "text": "#e9ecf2",
        "muted": "#b3b9c6",
    }

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure(
        "Card.TFrame",
        background=palette["card"],
        relief="ridge",
        borderwidth=1,
    )
    style.configure(
        "Heading.TLabel",
        background=palette["card"],
        foreground=palette["text"],
        font=("Inter", 14, "bold"),
    )
    style.configure(
        "Body.TLabel",
        background=palette["card"],
        foreground=palette["muted"],
        font=("Inter", 10),
    )
    style.configure(
        "TButton",
        font=("Inter", 10),
        foreground=palette["text"],
        background=palette["panel"],
        padding=8,
        borderwidth=0,
    )
    style.map(
        "TButton",
        background=[("active", palette["border"]), ("pressed", palette["border"])],
        foreground=[("disabled", palette["muted"])],
    )
    style.configure(
        "Accent.TButton",
        font=("Inter", 10, "bold"),
        foreground=palette["bg"],
        background=palette["accent"],
        padding=8,
        borderwidth=0,
        focusthickness=0,
    )
    style.map(
        "Accent.TButton",
        background=[("active", "#ffbe4c"), ("pressed", "#d89c2f")],
        foreground=[("disabled", palette["muted"])],
    )
    style.configure(
        "TEntry",
        fieldbackground=palette["panel"],
        background=palette["panel"],
        foreground=palette["text"],
        insertcolor=palette["text"],
        relief="flat",
        padding=6,
    )
    style.map(
        "TEntry",
        fieldbackground=[("focus", palette["panel"]), ("!focus", palette["panel"])],
    )

    content = ttk.Frame(root, padding=16, style="Card.TFrame")
    content.grid(row=0, column=0, padx=12, pady=12, sticky="nsew")

    title = ttk.Label(
        content,
        text="Форма выгрузки ответов",
        style="Heading.TLabel",
    )
    title.grid(row=0, column=0, columnspan=3, pady=(0, 14), sticky="w")

    input_dir_var = tk.StringVar()
    output_file_var = tk.StringVar(value=str((Path.cwd() / "combined.xlsx").resolve()))
    status_var = tk.StringVar(value="Заполните поля и нажмите «Собрать отчёт».")
    generate_pdf_var = tk.BooleanVar(value=False)

    def browse_input_dir() -> None:
        path = filedialog.askdirectory(title="Папка с JSON файлами")
        if path:
            input_dir_var.set(path)

    def browse_output_file() -> None:
        filename = filedialog.asksaveasfilename(
            title="Сохранить отчёт",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
            initialfile="combined.xlsx",
        )
        if filename:
            output_file_var.set(filename)

    def run_merge() -> None:
        input_dir = Path(input_dir_var.get()).expanduser()
        output_path = Path(output_file_var.get()).expanduser()
        generate_pdfs = bool(generate_pdf_var.get())

        if not input_dir.exists() or not input_dir.is_dir():
            messagebox.showerror("Ошибка", "Укажите существующую папку с JSON файлами.")
            return

        try:
            merged = merge_json_directory(input_dir, output_path, generate_pdfs=generate_pdfs)
        except Exception as exc:  # noqa: BLE001 - user-facing helper
            messagebox.showerror("Не удалось собрать отчёт", str(exc))
            status_var.set("Ошибка при сборке. Попробуйте снова.")
            return

        pdf_note = " и PDF" if generate_pdfs else ""
        status_var.set(f"Готово! Объединено файлов: {merged} → {output_path}{pdf_note}")
        messagebox.showinfo(
            "Готово",
            f"Сохранено: {output_path}\nФайлов: {merged}" + ("\nСозданы индивидуальные PDF" if generate_pdfs else ""),
        )

    # Поля формы
    ttk.Label(content, text="Папка с ответами", style="Body.TLabel").grid(
        row=1, column=0, sticky="w", pady=(0, 6)
    )
    input_entry = ttk.Entry(content, textvariable=input_dir_var, width=48)
    input_entry.grid(row=2, column=0, columnspan=2, sticky="we", padx=(0, 8))
    ttk.Button(content, text="Обзор", command=browse_input_dir).grid(row=2, column=2, sticky="we")

    ttk.Label(content, text="Файл отчёта", style="Body.TLabel").grid(
        row=3, column=0, sticky="w", pady=(12, 6)
    )
    output_entry = ttk.Entry(content, textvariable=output_file_var, width=48)
    output_entry.grid(row=4, column=0, columnspan=2, sticky="we", padx=(0, 8))
    ttk.Button(content, text="Сохранить как", command=browse_output_file).grid(
        row=4, column=2, sticky="we"
    )

    ttk.Checkbutton(
        content,
        text="Создавать PDF с детализацией баллов для каждого ответа",
        variable=generate_pdf_var,
        style="TCheckbutton",
    ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))

    action = ttk.Button(
        content,
        text="Собрать отчёт",
        style="Accent.TButton",
        command=run_merge,
    )
    action.grid(row=6, column=0, columnspan=3, pady=(16, 8), sticky="we")

    status_label = ttk.Label(content, textvariable=status_var, style="Body.TLabel", wraplength=420)
    status_label.grid(row=7, column=0, columnspan=3, sticky="w")

    for child in content.winfo_children():
        child.grid_configure(padx=4, pady=2)

    input_entry.focus_set()
    root.mainloop()


def main() -> None:
    if len(sys.argv) > 1:
        args = parse_args()
        merged = merge_json_directory(args.input_dir, args.output, generate_pdfs=args.generate_pdf)
        print(f"Merged {merged} JSON files into {args.output}")
    else:
        launch_gui()


if __name__ == "__main__":
    main()
